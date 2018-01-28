#!/usr/bin/env python3

import sys
import re
import argparse

from openpyxl import load_workbook
import xlwt

import html as html_lib
import urllib.request
import urllib.error


TITLE_PATTERN = re.compile('<title.*?>(.*?)</title>')
H1_PATTERN = re.compile('<h1.*?>(?:<[A-Za-z]+?(?:\s.*?)?>)*(.*?)(?:</.+?>)*?</h1>', re.DOTALL)


def getPageData(url):
    data = {}

    try:
        resp = urllib.request.urlopen(url)
        html = html_lib.unescape(resp.read().decode('utf-8'))
    except urllib.error.URLError as e:
        data['error'] = 'url: {}'.format(e.reason)
    except urllib.error.HTTPError as e:
        data['error'] = 'http: {}'.format(e.code)
    except Exception:
        data['error'] = 'unknown error'
    else:
        title = TITLE_PATTERN.search(html)
        h1 = H1_PATTERN.search(html)

        data['title'] = title[1].strip(' \r\t\n') if title else ''
        data['h1'] = h1[1].strip(' \r\t\n') if h1 else ''

    return data


def compare_urls(data, other_url, group_paths=None):
    offset = []

    base_url = data[0]['url'].rstrip('/')

    path_pattern = re.compile('^{}(/.*)'.format(base_url))

    passed_group_paths = set()
    for obj in data:
        rel_path = path_pattern.search(obj['url'])[1]

        if group_paths:
            is_passed = False
            for path in group_paths:
                if rel_path.startswith(path):
                    if path in passed_group_paths:
                        is_passed = True
                    else:
                        passed_group_paths.add(path)

                    break

            if is_passed:
                continue

        print('Читаем: {}'.format(rel_path))
        other_data = getPageData(other_url + rel_path)

        total = {}

        if 'error' not in other_data:
            title_eq = obj['title'] == other_data['title']
            h1_eq = obj['h1'] == other_data['h1']
            if title_eq and h1_eq:
                continue

            if not title_eq:
                total['exp_title'] = obj['title']
                total['act_title'] = other_data['title']
            if not h1_eq:
                total['exp_h1'] = obj['h1']
                total['act_h1'] = other_data['h1']
        else:
            total['error'] = other_data['error']

        total['url'] = rel_path

        offset.append(total)

    return offset


def parse_excel(path):
    result = []

    wb = load_workbook(path)
    sheet = wb.active

    for row in range(2, sheet.max_row+1):
        url = sheet.cell(row=row, column=1).value
        title = sheet.cell(row=row, column=2).value
        h1 = sheet.cell(row=row, column=3).value
        result.append({
            'url': url.strip(),
            'title': title.strip() if title else '',
            'h1': h1.strip() if h1 else ''
        })

    return result


def output_excel(data):
    wb = xlwt.Workbook(encoding='utf-8')
    sheet = wb.add_sheet('Sheet 1')

    sheet.write(0, 0, 'URL')
    sheet.write(0, 1, 'TITLE (Expected)')
    sheet.write(0, 2, 'TITLE (Actual)')
    sheet.write(0, 3, 'H1 (Expected)')
    sheet.write(0, 4, 'H1 (Actual)')
    sheet.write(0, 5, 'ERROR')

    line_num = 1

    fields = dict(zip(range(6), ('url', 'exp_title', 'act_title',
                                 'exp_h1', 'act_h1', 'error')))

    for record in data:
        row = sheet.row(line_num)
        line_num += 1

        for n, field in fields.items():
            if field in record:
                row.write(n, record[field])

    wb.save('report.xls')


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument(
        'excel_path',
        help='Абсолютный или относительный путь до таблицы с данными')
    parser.add_argument(
        'mirror_path',
        help='Базовый путь, с которым сравниваем урлы из таблицы')
    parser.add_argument(
        '-g',
        help=('Урлы, для которых достаточно проверить первую подкатегорию. '
              'Перечисление урлов необходимо выделять кавычками: "URL1, ..."')
    )
    args = parser.parse_args()

    print('Читаем таблицу...')
    table = parse_excel(args.excel_path)

    print('Начинаем чтение url-ов...')
    if args.g:
        paths = args.g.split()
        result = compare_urls(table, args.mirror_path, paths)
    else:
        result = compare_urls(table, args.mirror_path)

    print('Сохраняем результат...')
    output_excel(result)
